#!/usr/bin/env python3
"""
DuckDB Formula Evaluator Benchmark

Measures time and memory for Excel formula evaluation using DuckDB.
Uses xlsx2csv library for CSV conversion, DuckDB read_csv_auto for loading,
and xlsxwriter for writing output.

Compatible with JS and LibreOffice benchmark output format.
"""

import sys
import duckdb
import psutil
import time
import json
import threading
import re
from pathlib import Path
from openpyxl import load_workbook
from xlsx2csv import Xlsx2csv
import xlsxwriter
import tempfile
import os

# Import FormulaEvaluator for complex formulas
from lib.formula_evaluator import FormulaEvaluator


def parse_formula_pattern(formula: str) -> dict:
    """
    Parse a formula to detect patterns that can be vectorized.

    Returns:
        {
            'type': 'simple' | 'cross_sheet' | 'complex',
            'pattern': e.g., 'A+B' or 'A*2' or 'Sheet1!A',
            'columns': list of columns involved
        }
    """
    # Remove leading = and whitespace
    formula = formula.lstrip('=').strip()

    # Simple arithmetic: A2+B2, A2*2, etc.
    # Pattern: single letter column refs with operators
    simple_arithmetic = re.match(r'^([A-Z])\d+\s*([+\-*/])\s*([A-Z])\d+$', formula)
    if simple_arithmetic:
        return {
            'type': 'simple',
            'pattern': f'{simple_arithmetic.group(1)} {simple_arithmetic.group(2)} {simple_arithmetic.group(3)}',
            'columns': [simple_arithmetic.group(1), simple_arithmetic.group(3)]
        }

    # Simple scalar operation: A2*2, B2/10, etc.
    simple_scalar = re.match(r'^([A-Z])\d+\s*([+\-*/])\s*(\d+(?:\.\d+)?)$', formula)
    if simple_scalar:
        return {
            'type': 'simple',
            'pattern': f'{simple_scalar.group(1)} {simple_scalar.group(2)} {simple_scalar.group(3)}',
            'columns': [simple_scalar.group(1)]
        }

    # Cross-sheet reference: Sheet1!A2
    cross_sheet = re.match(r'^([A-Za-z0-9_]+)!([A-Z])\d+$', formula)
    if cross_sheet:
        return {
            'type': 'cross_sheet',
            'source_sheet': cross_sheet.group(1),
            'column': cross_sheet.group(2)
        }

    return {'type': 'complex'}


def measure_benchmark(n: str) -> dict:
    """Measure benchmark: time, baseline memory, peak memory, used memory."""

    # Build paths
    input_path = Path(f"/tmp/benchmark/test_{n}.xlsx")
    output_path = Path(f"/tmp/out_duckdb_{n}.xlsx")

    # Extract row count for reporting
    if n == "max":
        rows_report = 1048576
    elif n.startswith("2sheet_"):
        rows_report = int(n.replace("2sheet_", ""))
    else:
        rows_report = int(n)

    process = psutil.Process()

    # Measure baseline
    baselineMB = process.memory_info().rss / 1024 / 1024
    peakMB = baselineMB

    measuring = False

    def memory_monitor():
        nonlocal peakMB
        while measuring:
            peakMB = max(peakMB, process.memory_info().rss / 1024 / 1024)
            time.sleep(0.01)

    # Start memory monitor BEFORE any work
    measuring = True
    monitor = threading.Thread(target=memory_monitor, daemon=True)
    monitor.start()

    # Measure time
    start = time.time()

    # Step 1: Convert XLSX to CSV using xlsx2csv library
    temp_dir = tempfile.mkdtemp()
    csv_files = {}

    wb = load_workbook(input_path, data_only=False)

    for sheet_name in wb.sheetnames:
        csv_path = os.path.join(temp_dir, f"{sheet_name}.csv")
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            converter = Xlsx2csv(input_path, sheetname=sheet_name, outputfile=f)
            converter.convert()
        csv_files[sheet_name] = csv_path

    # Step 2: Load CSVs into DuckDB using read_csv_auto
    conn = duckdb.connect(':memory:')

    for sheet_name, csv_path in csv_files.items():
        table_name = sheet_name.lower().replace(' ', '_')
        conn.execute(f"""
            CREATE TABLE {table_name} AS
            SELECT * FROM read_csv_auto('{csv_path}', header=False)
        """)

    # Step 3: Read formulas from Excel and evaluate
    evaluator = FormulaEvaluator(conn)
    ws = wb.active

    # Collect formulas by column for batch processing
    formulas_by_column = {}

    for col in range(1, ws.max_column + 1):
        sample_cell = ws.cell(2, col) if ws.max_row >= 2 else None
        if not sample_cell or sample_cell.data_type != 'f':
            continue

        pattern = parse_formula_pattern(sample_cell.value)
        formulas_by_column[col] = {
            'pattern': pattern,
            'formula': sample_cell.value
        }

    # Evaluate formulas
    for col, info in formulas_by_column.items():
        pattern = info['pattern']
        formula = info['formula']
        table_name = ws.title.lower().replace(' ', '_')

        if pattern['type'] in ('simple', 'cross_sheet'):
            # Vectorized evaluation using SQL
            if pattern['type'] == 'simple':
                parts = pattern['pattern'].split()
                col1 = f"col{ord(parts[0]) - ord('A')}"
                op = parts[1]
                col2_or_val = parts[2]

                if col2_or_val.isalpha():
                    col2 = f"col{ord(col2_or_val) - ord('A')}"
                    expr = f'"{col1}" {op} "{col2}"'
                else:
                    expr = f'"{col1}" {op} {col2_or_val}'

                conn.execute(f'ALTER TABLE {table_name} ADD COLUMN _result{col} DOUBLE')
                conn.execute(f'UPDATE {table_name} SET _result{col} = {expr}')

        else:
            # Complex formulas - evaluate using FormulaEvaluator
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col)
                if cell.data_type == 'f':
                    row_ctx = {}
                    for ctx_col in range(1, ws.max_column + 1):
                        ctx_cell = ws.cell(row, ctx_col)
                        if ctx_cell.data_type != 'f':
                            col_letter = chr(ord('A') + ctx_col - 1)
                            row_ctx[f"{col_letter}{row}"] = ctx_cell.value
                    try:
                        sql = evaluator.excel_to_sql(formula, ws.title, row_ctx)
                        result = conn.execute(sql).fetchone()[0]
                    except Exception:
                        result = None

    # Step 4: Write results to XLSX using xlsxwriter
    workbook = xlsxwriter.Workbook(output_path)

    for sheet_name in wb.sheetnames:
        table_name = sheet_name.lower().replace(' ', '_')

        # Get data from DuckDB
        result_df = conn.execute(f"SELECT * FROM {table_name}").fetchdf()

        # Write to worksheet
        worksheet = workbook.add_worksheet(sheet_name)

        for row_idx, row_data in enumerate(result_df.itertuples(index=False)):
            for col_idx, value in enumerate(row_data):
                worksheet.write(row_idx, col_idx, value)

    workbook.close()

    # Cleanup
    end = time.time()
    measuring = False
    monitor.join(timeout=1)

    # Clean up temp files
    for csv_path in csv_files.values():
        if os.path.exists(csv_path):
            os.remove(csv_path)
    os.rmdir(temp_dir)

    timeSeconds = end - start
    peakTotalMB = peakMB
    usedMB = peakMB - baselineMB

    return {
        "rows": rows_report,
        "peakTotalMB": round(peakTotalMB, 1),
        "usedMB": round(usedMB, 1),
        "baselineMB": round(baselineMB, 1),
        "timeSeconds": round(timeSeconds, 3)
    }


def main():
    n = sys.argv[1] if len(sys.argv) > 1 else "10000"
    result = measure_benchmark(n)
    print(json.dumps(result))

if __name__ == "__main__":
    main()
